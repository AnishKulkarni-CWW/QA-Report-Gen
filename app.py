"""
QA Work-Hours Analytics Dashboard
----------------------------------
A single-page Streamlit dashboard for analyzing QA team work-hour trackers.

Upload the QA tracker workbook -> instant KPI summary + charts (line trend,
daily intensity grid, QA comparison bars, hours mix, searchable daily log) ->
filter by QA / Year / Month / Week / Day (multi-select + calendar) -> click
"Prepare Export" once to build Excel/PDF snapshots of exactly what's on screen.

Run:  streamlit run app.py
"""

import io
import re
import warnings
from datetime import datetime, date

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

warnings.filterwarnings("ignore")

# ============================================================================
# PAGE CONFIG & GLOBAL STYLE
# ============================================================================
st.set_page_config(
    page_title="QA Hours Dashboard",
    page_icon="\U0001F4CA",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---- Palette (matches the reference screenshot: dark navy header, teal/slate accents) ----
INK = "#0F1729"           # dark navy header background
INK_TEXT = "#E7EAF3"
PAGE_BG = "#F4F5F8"
CARD_BG = "#FFFFFF"
BORDER = "#E7E9F0"
TEXT_MAIN = "#1F2333"
TEXT_MUTED = "#7A7F91"

BILLABLE_COLOR = "#2F9E8F"     # teal
NONBILL_COLOR = "#E0A72E"      # amber
NOTWORKED_COLOR = "#C1543D"    # brick red
ACCENT = "#2F9E8F"


# Note: intentionally does NOT reuse BILLABLE_COLOR (#2F9E8F) or
# NOTWORKED_COLOR (#C1543D) — those two hues are reserved for the
# Billable/Non-Billable/Not-Worked state encoding in the stacked and donut
# charts, so a QA-identity color here can never be mistaken for that meaning.
QA_PALETTE = ["#4C6FA6", "#D97A46", "#6E7FC9", "#5B9279",
              "#8E9257", "#DCB13A", "#A65D8C", "#9C6FA6"]

CUSTOM_CSS = f"""
<style>
    .stApp {{
        background: {PAGE_BG};
    }}
    #MainMenu {{visibility: hidden;}}
    footer {{visibility: hidden;}}
    header[data-testid="stHeader"] {{
        background: transparent;
        height: 3rem;
    }}
    .block-container {{
        padding-top: 1rem !important;
        padding-bottom: 3rem !important;
        max-width: 1400px;
    }}

    /* Top banner mimicking the reference design */
    .top-banner {{
        background: {INK};
        border-radius: 16px;
        padding: 18px 28px;
        margin-bottom: 20px;
        color: {INK_TEXT};
    }}
    .top-banner .eyebrow {{
        font-size: 0.7rem;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        color: #8B93AD;
        font-weight: 600;
        margin-bottom: 2px;
    }}
    .top-banner .title {{
        font-size: 1.5rem;
        font-weight: 800;
        color: #FFFFFF;
    }}

    .kpi-card {{
        background: {CARD_BG};
        border-radius: 14px;
        padding: 16px 18px 14px 18px;
        box-shadow: 0 1px 3px rgba(15,23,41,0.06);
        border: 1px solid {BORDER};
        min-height: 108px;
        overflow: visible;
        margin-bottom: 4px;
    }}
    .kpi-label {{
        font-size: 0.72rem;
        font-weight: 700;
        color: {TEXT_MUTED};
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 6px;
    }}
    .kpi-value {{
        font-size: 1.8rem;
        font-weight: 800;
        color: {TEXT_MAIN};
        line-height: 1.1;
    }}
    .kpi-value .unit {{
        font-size: 0.95rem;
        font-weight: 600;
        color: {TEXT_MUTED};
        margin-left: 3px;
    }}
    .kpi-sub {{
        font-size: 0.74rem;
        color: {TEXT_MUTED};
        margin-top: 4px;
    }}

    .panel {{
        background: {CARD_BG};
        border-radius: 16px;
        padding: 20px 22px;
        box-shadow: 0 1px 3px rgba(15,23,41,0.06);
        border: 1px solid {BORDER};
        margin-bottom: 20px;
    }}
    .panel-title {{
        font-size: 1.05rem;
        font-weight: 800;
        color: {TEXT_MAIN};
    }}
    .panel-sub {{
        font-size: 0.78rem;
        color: {TEXT_MUTED};
    }}

    .qa-card {{
        background: {CARD_BG};
        border-radius: 14px;
        padding: 14px 16px;
        box-shadow: 0 1px 3px rgba(15,23,41,0.06);
        border: 1px solid {BORDER};
        margin-bottom: 12px;
    }}
    .qa-name {{
        font-size: 1.0rem;
        font-weight: 800;
        color: {TEXT_MAIN};
    }}

    div[data-testid="stFileUploader"] {{
        background: {CARD_BG};
        border-radius: 14px;
        padding: 12px;
        border: 1.5px dashed #C9CEDD;
    }}

    .modebar {{ display: none !important; }}

    /* Sidebar */
    section[data-testid="stSidebar"] {{
        background: {CARD_BG};
        border-right: 1px solid {BORDER};
    }}

    .stButton > button {{
        border-radius: 10px;
        font-weight: 700;
    }}
    div[data-testid="stDownloadButton"] > button {{
        border-radius: 10px;
        font-weight: 700;
        background: {INK};
        color: white;
        border: none;
    }}

    /* Give multiselect (QA / Years / Months / Weeks) and date-input dropdowns
       a visible rounded border so they read as proper closed input fields. */
    div[data-baseweb="select"] > div {{
        border: 1.5px solid {BORDER} !important;
        border-radius: 10px !important;
        box-shadow: none !important;
    }}
    div[data-baseweb="select"] > div:focus-within {{
        border-color: {ACCENT} !important;
    }}
    div[data-testid="stDateInput"] > div {{
        border: 1.5px solid {BORDER} !important;
        border-radius: 10px !important;
    }}
    div[data-testid="stDateInput"] input {{
        border: none !important;
    }}

    /* Breathing room after the KPI summary row, before the next panel/card */
    .kpi-row-spacer {{
        height: 20px;
    }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Disable Plotly's floating toolbar everywhere (prevents accidental clicks).
PLOTLY_CONFIG = {"displayModeBar": False, "displaylogo": False, "scrollZoom": False}

TODAY = pd.Timestamp(datetime.now().date())

# ============================================================================
# DATA LOADING & CLEANING
# ============================================================================

HEADER_ALIASES = {
    "enter qa name here": "QA Name",
    "qa name": "QA Name",
    "date": "Date",
    "day": "Day",
    "month": "Month",
    "billable hours": "Billable Hours",
    "non-billable hours": "Non-Billable Hours",
    "non billable hours": "Non-Billable Hours",
    "hours not worked": "Hours Not Worked",
    "total hours": "Total Hours",
    "project name": "Project Name",
    "comment": "Comment",
    "comments": "Comment",
}

SKIP_SHEETS = {"master data", "config", "instructions"}
WEEKEND_TOKENS = {"sat", "sun", "saturday", "sunday"}


def _norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _find_header_row(raw: pd.DataFrame, max_scan: int = 5):
    for i in range(min(max_scan, len(raw))):
        row_vals = [_norm(v) for v in raw.iloc[i].tolist()]
        hits = sum(1 for v in row_vals if v in HEADER_ALIASES)
        if hits >= 3:
            return i
    return 0


def _to_number(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)
    s = str(x).strip()
    if s == "" or _norm(s) in WEEKEND_TOKENS:
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan


def parse_workbook(file) -> pd.DataFrame:
    """Read every per-QA sheet in the workbook and return one tidy dataframe.
    Rows dated after today are dropped, even if present in the source file."""
    xls = pd.ExcelFile(file)
    frames = []

    for sheet in xls.sheet_names:
        if _norm(sheet) in SKIP_SHEETS:
            continue

        raw = xls.parse(sheet, header=None)
        if raw.empty or raw.shape[0] < 2:
            continue

        header_row = _find_header_row(raw)
        header_vals = raw.iloc[header_row].tolist()
        col_map = {}
        for idx, v in enumerate(header_vals):
            key = _norm(v)
            if key in HEADER_ALIASES:
                col_map[idx] = HEADER_ALIASES[key]

        if "Billable Hours" not in col_map.values():
            continue

        body = raw.iloc[header_row + 1:].copy()
        body = body.rename(columns=col_map)
        keep_cols = [c for c in body.columns if isinstance(c, str) and c in HEADER_ALIASES.values()]
        body = body[keep_cols]
        body = body.loc[:, ~body.columns.duplicated()]

        if "QA Name" not in body.columns:
            body["QA Name"] = sheet
        body["QA Name"] = body["QA Name"].fillna(sheet)
        body.loc[body["QA Name"].astype(str).str.strip() == "", "QA Name"] = sheet

        if "Date" in body.columns:
            body["Date"] = pd.to_datetime(body["Date"], errors="coerce")
        else:
            body["Date"] = pd.NaT

        body = body[body["Date"].notna()]
        if body.empty:
            continue

        for c in ["Billable Hours", "Non-Billable Hours", "Hours Not Worked", "Total Hours"]:
            if c in body.columns:
                body[c] = body[c].apply(_to_number)
            else:
                body[c] = np.nan

        hour_cols = ["Billable Hours", "Non-Billable Hours", "Hours Not Worked", "Total Hours"]
        body = body[~body[hour_cols].isna().all(axis=1)]
        if body.empty:
            continue

        body[hour_cols] = body[hour_cols].fillna(0.0)

        computed_total = body["Billable Hours"] + body["Non-Billable Hours"] + body["Hours Not Worked"]
        body["Total Hours"] = np.where(
            (body["Total Hours"] <= 0) | (body["Total Hours"].isna()),
            computed_total,
            body["Total Hours"],
        )

        body["Day"] = body["Date"].dt.day_name().str.slice(0, 3)
        body["Month"] = body["Date"].dt.strftime("%b")
        body["Year"] = body["Date"].dt.year

        if "Comment" in body.columns:
            body["Comment"] = body["Comment"].fillna("").astype(str).str.strip()
        else:
            body["Comment"] = ""

        body["QA Name"] = body["QA Name"].astype(str).str.strip()
        body["QA Name"] = body["QA Name"].str.replace(r"[._]+", " ", regex=True)
        body["QA Name"] = body["QA Name"].str.replace(r"\s+", " ", regex=True).str.strip()
        body["QA Name"] = body["QA Name"].str.title()

        final = body[["QA Name", "Date", "Day", "Month", "Year",
                      "Billable Hours", "Non-Billable Hours",
                      "Hours Not Worked", "Total Hours", "Comment"]].copy()
        final = final[final["QA Name"] != ""]
        frames.append(final)

    if not frames:
        return pd.DataFrame(columns=["QA Name", "Date", "Day", "Month", "Year",
                                      "Billable Hours", "Non-Billable Hours",
                                      "Hours Not Worked", "Total Hours", "Comment"])

    out = pd.concat(frames, ignore_index=True)
    out = out.drop_duplicates(subset=["QA Name", "Date"], keep="first")

    # Hard rule: never show data beyond today, even if the workbook has it.
    out = out[out["Date"] <= TODAY]

    out = out.sort_values(["QA Name", "Date"]).reset_index(drop=True)
    return out


# ============================================================================
# CHART BUILDERS
# ============================================================================

def qa_color_map(qa_names):
    return {name: QA_PALETTE[i % len(QA_PALETTE)] for i, name in enumerate(sorted(qa_names))}


def _teal_scale(t):
    """Map t in [0, 1] to a hex color on a light-to-dark teal ramp (same
    family as BILLABLE_COLOR). Used for value-encoded bars where color
    represents magnitude, not category identity."""
    t = max(0.0, min(1.0, t))
    light = (0xD3, 0xEE, 0xEA)   # pale teal
    dark = (0x17, 0x5E, 0x54)    # deep teal
    rgb = tuple(int(light[i] + (dark[i] - light[i]) * t) for i in range(3))
    return f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def donut_chart(billable, nonbill, notworked, title="Team Utilization Split"):
    labels = ["Billable", "Non-Billable", "Not Worked"]
    values = [billable, nonbill, notworked]
    colors = [BILLABLE_COLOR, NONBILL_COLOR, NOTWORKED_COLOR]
    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.62,
        marker=dict(colors=colors, line=dict(color="#FFFFFF", width=3)),
        textinfo="percent", textfont=dict(size=13, color="white", family="Arial Black"),
        hovertemplate="%{label}: %{value:.1f} hrs (%{percent})<extra></extra>",
        sort=False,
    )])
    total = billable + nonbill + notworked
    util = (billable / total * 100) if total else 0
    fig.update_layout(
        title=dict(text=title, font=dict(size=15, color=TEXT_MAIN), x=0.0, xanchor="left"),
        annotations=[dict(text=f"<b>{util:.0f}%</b><br><span style='font-size:11px;color:#7A7F91'>Utilization</span>",
                           x=0.5, y=0.5, showarrow=False, font=dict(size=20, color=TEXT_MAIN))],
        showlegend=True,
        legend=dict(orientation="h", yanchor="top", y=-0.08, xanchor="center", x=0.5),
        margin=dict(t=50, b=55, l=10, r=10),
        height=360,
        paper_bgcolor="rgba(0,0,0,0)",
    )
    return fig


def bar_chart_by_qa(df_period, title="QA Comparison — Total Hours"):
    # Canonical sort used across the whole app: Total Hours, descending.
    # Kept as ascending=True here on purpose — Plotly draws horizontal bars
    # bottom-to-top, so an ascending sort on the underlying data renders the
    # highest total at the TOP of the chart, i.e. visually descending.
    grp = df_period.groupby("QA Name")["Total Hours"].sum().reset_index()
    grp = grp.sort_values("Total Hours", ascending=True)

    # Single-hue value scale (teal, matching BILLABLE_COLOR's family) instead
    # of the categorical QA_PALETTE. Color here doesn't encode QA identity —
    # it encodes magnitude — so a continuous scale reads more honestly than a
    # rainbow of unrelated per-person colors. This also keeps the
    # teal/amber/red trio reserved strictly for Billable/Non-Billable/Not-Worked
    # in the stacked and donut charts, with no accidental reuse here.
    max_hours = grp["Total Hours"].max() if len(grp) else 0
    bar_colors = [
        _teal_scale(v / max_hours if max_hours else 0) for v in grp["Total Hours"]
    ]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=grp["QA Name"], x=grp["Total Hours"], orientation="h",
        marker_color=bar_colors,
        text=[f"{v:,.1f}" for v in grp["Total Hours"]],
        textposition="outside",
        hovertemplate="%{y}: %{x:.1f} hrs<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text=title, font=dict(size=15, color=TEXT_MAIN), x=0.0, xanchor="left"),
        xaxis=dict(title="Hours", gridcolor="#F0F1F5"),
        yaxis=dict(title=""),
        showlegend=False,
        margin=dict(t=50, b=20, l=10, r=40),
        height=max(320, 42 * len(grp) + 90),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig


def hours_mix_chart(df_period, title="Hours Mix — Billable / Non-Billable / Not Worked"):
    # Canonical sort: Total Hours, descending — same rule used by the bar
    # chart, the breakdown cards, and the summary table below.
    grp = df_period.groupby("QA Name")[
        ["Billable Hours", "Non-Billable Hours", "Hours Not Worked"]
    ].sum().reset_index()
    grp["Total Hours"] = grp["Billable Hours"] + grp["Non-Billable Hours"] + grp["Hours Not Worked"]
    grp = grp.sort_values("Total Hours", ascending=False)

    fig = go.Figure()
    fig.add_trace(go.Bar(x=grp["QA Name"], y=grp["Billable Hours"], name="Billable", marker_color=BILLABLE_COLOR,
                          hovertemplate="Billable: %{y:.1f} hrs<extra></extra>"))
    fig.add_trace(go.Bar(x=grp["QA Name"], y=grp["Non-Billable Hours"], name="Non-Billable", marker_color=NONBILL_COLOR,
                          hovertemplate="Non-Billable: %{y:.1f} hrs<extra></extra>"))
    fig.add_trace(go.Bar(x=grp["QA Name"], y=grp["Hours Not Worked"], name="Not Worked", marker_color=NOTWORKED_COLOR,
                          hovertemplate="Not Worked: %{y:.1f} hrs<extra></extra>"))
    fig.update_layout(
        barmode="stack",
        title=dict(text=title, font=dict(size=15, color=TEXT_MAIN), x=0.0, xanchor="left"),
        xaxis=dict(title="", gridcolor="#F0F1F5"),
        yaxis=dict(title="Hours", gridcolor="#F0F1F5"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
        margin=dict(t=50, b=10, l=10, r=10),
        height=420,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig


def qa_mini_donut(row, qa_name):
    b, n, w = row["Billable Hours"], row["Non-Billable Hours"], row["Hours Not Worked"]
    fig = go.Figure(data=[go.Pie(
        labels=["Billable", "Non-Billable", "Not Worked"], values=[b, n, w], hole=0.65,
        marker=dict(colors=[BILLABLE_COLOR, NONBILL_COLOR, NOTWORKED_COLOR], line=dict(color="white", width=2)),
        # Show each slice's own share as an on-chart label (matching the big
        # donut's textinfo="percent") so the split is readable at a glance,
        # not only on hover.
        textinfo="percent",
        textposition="inside",
        textfont=dict(size=10, color="white", family="Arial Black"),
        hovertemplate="%{label}: %{value:.1f} hrs (%{percent})<extra></extra>",
        sort=False,
    )])
    total = b + n + w
    util = (b / total * 100) if total else 0
    fig.update_layout(
        annotations=[dict(text=f"<b>{util:.0f}%</b><br><span style='font-size:8px;color:#7A7F91'>Utilization</span>",
                           x=0.5, y=0.5, showarrow=False,
                           font=dict(size=15, color=TEXT_MAIN))],
        showlegend=False, margin=dict(t=0, b=0, l=0, r=0), height=150,
        paper_bgcolor="rgba(0,0,0,0)",
    )
    return fig


# ============================================================================
# EXPORT HELPERS  (only ever called after the user clicks "Prepare Export")
# ============================================================================

def _safe_filename_fragment(text, max_len=60):
    """Strip characters that are unsafe in Windows/Mac filenames (plus
    parentheses, which are legal but look messy in a filename) and collapse
    whitespace/commas into underscores, so export filenames never break when
    downloaded on either OS. Truncates long fragments so the full filename
    stays reasonable."""
    text = re.sub(r'[\\/:*?"<>|()]', "", text)
    text = re.sub(r"[\s,]+", "_", text.strip())
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")
    return text[:max_len] if len(text) > max_len else text


def _build_export_filename_base(period_label, sel_qas, all_qas):
    """Build the shared filename base (without extension) for an export,
    reflecting both the time period AND the QA selection at the moment
    "Prepare Exports" was clicked -- so a file filtered to 2 of 8 QAs is
    identifiable from its filename alone, not just its contents."""
    period_part = _safe_filename_fragment(period_label)

    if not sel_qas or set(sel_qas) == set(all_qas):
        # No QA-specific suffix needed: either everyone is included (the
        # default, most common case) or nothing is (period_label alone,
        # e.g. "No months selected", already says enough).
        qa_part = ""
    elif len(sel_qas) <= 3:
        # A handful of named QAs: spell them out, it's still a short, readable filename.
        qa_part = "_" + "_".join(_safe_filename_fragment(q, max_len=20) for q in sorted(sel_qas))
    else:
        # Many QAs but not all: naming them all would make the filename
        # unwieldy, so summarize by count instead.
        qa_part = f"_{len(sel_qas)}QAs"

    base = f"QA_Dashboard_{period_part}{qa_part}"
    return base if base else "QA_Dashboard_Export"


KALEIDO_PROBE_TIMEOUT_SECONDS = 8   # one-time-per-session viability check (see _check_kaleido_available)
KALEIDO_TIMEOUT_SECONDS = 10        # per real figure render; called up to ~22 times in one export cycle,
                                     # so this is a multiplicative cost, not a fixed one -- kept tight but
                                     # still generous given the render resolutions were also cut ~5x below.


def _run_with_timeout(fn, timeout_seconds, *args, **kwargs):
    """Run fn(*args, **kwargs) in a worker thread and enforce a hard wall-clock
    timeout. Needed because a broken kaleido install (most commonly: it can't
    find/launch a Chrome/Chromium runtime) doesn't always raise an exception —
    it can hang the underlying subprocess call indefinitely instead. A plain
    try/except never catches a hang, so this is the only reliable guard.
    Uses a thread (not signal.alarm) so it also works on Windows, since this
    app is meant to be packaged into a Windows .exe / Mac .dmg later.

    IMPORTANT: uses shutdown(wait=False) — if we let the ThreadPoolExecutor's
    context manager do its default wait=True shutdown, it blocks until the
    hung worker thread actually finishes, defeating the entire point of the
    timeout (the caller would still wait the full hang duration). The
    abandoned worker thread is left to die in the background (or run out
    the process's lifetime); Python threads don't get force-killed, but the
    caller gets control back at the deadline either way, which is what
    actually matters here."""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError

    pool = ThreadPoolExecutor(max_workers=1)
    future = pool.submit(fn, *args, **kwargs)
    try:
        result = future.result(timeout=timeout_seconds)
        pool.shutdown(wait=False)
        return result
    except FutureTimeoutError:
        pool.shutdown(wait=False)
        raise TimeoutError(
            f"Chart image rendering did not finish within {timeout_seconds} seconds "
            "and was abandoned. This almost always means kaleido's headless "
            "Chrome/Chromium subprocess is hanging or failing to launch in this "
            "environment, not that your computer is slow."
        )
    except Exception:
        pool.shutdown(wait=False)
        raise


def _check_kaleido_available():
    """Kaleido (==0.2.1, see requirements.txt) is required to turn Plotly
    figures into static images for the Excel/PDF exports.

    We deliberately use kaleido 0.2.1, NOT the newer 1.x line. 0.2.1 ships its
    OWN Chromium inside the wheel, so it renders fully offline with no separate
    Chrome install, no network download, and no "Install Chrome" step. That is
    exactly what lets the app be frozen into a self-contained Windows .exe /
    Mac app: PyInstaller can bundle kaleido's Chromium alongside the rest, and
    the end user needs nothing else on their machine. (kaleido 1.x removed the
    bundled browser and instead depends on a system Chrome at run time, which
    is fragile to ship in a frozen executable — hence the pin back to 0.2.1.)

    A successful `import kaleido` still doesn't 100% guarantee fig.to_image()
    works in every environment, so this runs a real, tiny, TIME-LIMITED
    end-to-end render to catch a broken install quickly instead of silently
    producing blank chart pages.

    Returns (ok: bool, message: str | None, chrome_missing: bool). Because
    0.2.1 has no separate Chrome dependency, chrome_missing is ALWAYS False —
    there is no external browser to install — so the caller's optional
    "Install Chrome" button (a kaleido-1.x-only affordance) never triggers."""
    try:
        import kaleido  # noqa: F401
    except ImportError:
        return False, (
            "The 'kaleido' package is not installed, so charts cannot be "
            "rendered into the Excel/PDF exports. Install it with:\n\n"
            "    pip install kaleido==0.2.1\n\n"
            "then restart the app and click Prepare Exports again."
        ), False

    def _probe():
        probe_fig = go.Figure(data=[go.Bar(x=[1], y=[1])])
        png_bytes = probe_fig.to_image(format="png", width=100, height=100, scale=1)
        if not png_bytes or len(png_bytes) < 200 or not png_bytes.startswith(b"\x89PNG"):
            raise RuntimeError("kaleido produced invalid/empty image data")
        return png_bytes

    try:
        _run_with_timeout(_probe, KALEIDO_PROBE_TIMEOUT_SECONDS)
        return True, None, False
    except Exception as e:
        # With 0.2.1 there is no "Chrome not installed" case (Chromium is
        # bundled), so we never flag chrome_missing. Any failure here is an
        # environment/version-mismatch issue, reported as such.
        return False, (
            f"Kaleido failed a quick test export ({e}). This build pins "
            "kaleido==0.2.1 and plotly==6.1.1, a verified-compatible pair that "
            "bundles its own Chromium and needs no separate Chrome install. "
            "If you see this, your installed versions likely drifted \u2014 "
            "reinstall the pinned versions:\n\n"
            "    pip install \"kaleido==0.2.1\" \"plotly==6.1.1\"\n\n"
            "then fully restart the app (not just refresh the browser) and "
            "click Prepare Exports again."
        ), False



def _fig_to_png_bytes(fig, width=900, height=500, scale=2):
    """Render a Plotly figure to PNG bytes, with a hard timeout. Raises on
    failure, on a timeout, OR on a suspiciously tiny/invalid result (instead
    of silently returning None, hanging forever, or trusting corrupt bytes)
    so export problems are always visible, never baked into a broken-looking
    PDF/Excel file with empty chart slots or an unresponsive UI."""
    def _render():
        return fig.to_image(format="png", width=width, height=height, scale=scale)

    png_bytes = _run_with_timeout(_render, KALEIDO_TIMEOUT_SECONDS)
    if not png_bytes or len(png_bytes) < 500 or not png_bytes.startswith(b"\x89PNG"):
        raise RuntimeError(
            "Chart image rendering returned invalid/empty PNG data — "
            "kaleido may be installed but not working correctly in this "
            "environment (common cause: missing Chrome/Chromium runtime "
            "that kaleido depends on)."
        )
    return png_bytes


# ============================================================================
# EXPORT CHART RENDERERS (matplotlib)
# ----------------------------------------------------------------------------
# The STATIC chart PNGs embedded in the Excel/PDF exports are rendered here
# with matplotlib, NOT with kaleido/Plotly. This is deliberate: matplotlib has
# no headless-browser/Chromium dependency, renders instantly with no
# subprocess or timeout, behaves identically on a laptop, a server, or a
# frozen PyInstaller .exe, and needs nothing installed beyond the Python
# package. kaleido (either the 1.x line that needs a system Chrome, or the
# 0.2.1 line whose bundled Chromium hangs on some machines) was the single
# source of "the PDF has no charts" problems, so the export path no longer
# touches it at all.
#
# The on-screen interactive dashboard still uses the Plotly figures
# (donut_chart / bar_chart_by_qa / hours_mix_chart / qa_mini_donut) exactly as
# before — only the exported images come from these matplotlib twins, drawn
# from the same aggregated data and the same palette so the PDF/Excel look
# like the live dashboard.
# ============================================================================
def _mpl():
    """Import matplotlib lazily (keeps app startup light) and return its
    pyplot + Patch handles. Agg backend = headless, safe everywhere."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch
    return plt, Patch


def _mpl_fig_png(fig, plt, dpi=200):
    b = io.BytesIO()
    fig.savefig(b, format="png", dpi=dpi, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    b.seek(0)
    return b.getvalue()


def mpl_team_donut(billable, nonbill, notworked):
    """Big team-utilization donut: three slices + center utilization %, with a
    compact legend baked in so the image is self-contained."""
    import math
    plt, Patch = _mpl()
    vals = [max(0.0, billable), max(0.0, nonbill), max(0.0, notworked)]
    total = sum(vals) or 1.0
    util = billable / total * 100.0

    fig, ax = plt.subplots(figsize=(4.4, 4.6))
    wedges, _ = ax.pie(
        vals, colors=[BILLABLE_COLOR, NONBILL_COLOR, NOTWORKED_COLOR],
        startangle=90, counterclock=False,
        wedgeprops=dict(width=0.38, edgecolor="white", linewidth=3),
    )
    for w, v in zip(wedges, vals):
        ang = (w.theta2 + w.theta1) / 2.0
        pct = v / total * 100.0
        if pct >= 4:
            ax.text(0.80 * math.cos(math.radians(ang)),
                    0.80 * math.sin(math.radians(ang)),
                    f"{pct:.1f}%", ha="center", va="center",
                    color="white", fontsize=11, fontweight="bold")
    ax.text(0, 0.08, f"{util:.0f}%", ha="center", va="center",
            fontsize=30, fontweight="bold", color=TEXT_MAIN)
    ax.text(0, -0.16, "Utilization", ha="center", va="center",
            fontsize=11, color=TEXT_MUTED)
    ax.set_aspect("equal")
    handles = [Patch(facecolor=BILLABLE_COLOR, label="Billable"),
               Patch(facecolor=NONBILL_COLOR, label="Non-Billable"),
               Patch(facecolor=NOTWORKED_COLOR, label="Not Worked")]
    ax.legend(handles=handles, loc="lower center", ncol=3,
              bbox_to_anchor=(0.5, -0.10), frameon=False, fontsize=10,
              handlelength=1.1, handleheight=1.1, columnspacing=1.4)
    fig.subplots_adjust(top=0.98, bottom=0.06, left=0.02, right=0.98)
    return _mpl_fig_png(fig, plt)


def mpl_bar_total(summary_df):
    """Horizontal Total-Hours-per-QA bars, highest at top, teal magnitude
    scale, value labels at the ends."""
    plt, _ = _mpl()
    grp = summary_df[["QA Name", "Total Hours"]].copy().sort_values("Total Hours", ascending=True)
    names = grp["QA Name"].tolist()
    vals = grp["Total Hours"].tolist()
    mx = max(vals) if vals else 0
    colors = [_teal_scale(v / mx if mx else 0) for v in vals]

    h = max(2.5, 0.26 * len(names) + 0.9)
    fig, ax = plt.subplots(figsize=(12.0, h))
    bars = ax.barh(names, vals, color=colors, edgecolor="none", height=0.68)
    ax.set_xlabel("Hours", fontsize=10, color=TEXT_MUTED)
    ax.grid(axis="x", color=BORDER, linewidth=0.8)
    ax.set_axisbelow(True)
    for s in ["top", "right", "left"]:
        ax.spines[s].set_visible(False)
    ax.spines["bottom"].set_color(BORDER)
    ax.tick_params(axis="y", length=0, labelsize=10)
    ax.tick_params(axis="x", length=0, labelsize=9, colors=TEXT_MUTED)
    pad = (mx * 0.02) if mx else 0.5
    for b, v in zip(bars, vals):
        ax.text(b.get_width() + pad, b.get_y() + b.get_height() / 2,
                f"{v:,.1f}", va="center", ha="left", fontsize=9,
                color=TEXT_MAIN, fontweight="bold")
    ax.set_xlim(0, (mx * 1.14) if mx else 1)
    fig.subplots_adjust(left=0.02, right=0.98, top=0.98, bottom=0.12)
    return _mpl_fig_png(fig, plt)


def mpl_hours_mix(summary_df):
    """Stacked vertical bars per QA (Billable / Non-Billable / Not Worked),
    sorted by total hours descending, legend beneath."""
    plt, _ = _mpl()
    grp = summary_df[["QA Name", "Billable Hours", "Non-Billable Hours", "Hours Not Worked"]].copy()
    grp["__t"] = grp["Billable Hours"] + grp["Non-Billable Hours"] + grp["Hours Not Worked"]
    grp = grp.sort_values("__t", ascending=False)
    names = grp["QA Name"].tolist()
    bill = grp["Billable Hours"].tolist()
    non = grp["Non-Billable Hours"].tolist()
    notw = grp["Hours Not Worked"].tolist()

    fig, ax = plt.subplots(figsize=(12.0, 2.95))
    x = list(range(len(names)))
    ax.bar(x, bill, color=BILLABLE_COLOR, label="Billable", width=0.62)
    ax.bar(x, non, bottom=bill, color=NONBILL_COLOR, label="Non-Billable", width=0.62)
    bottom2 = [b + n for b, n in zip(bill, non)]
    ax.bar(x, notw, bottom=bottom2, color=NOTWORKED_COLOR, label="Not Worked", width=0.62)
    ax.set_ylabel("Hours", fontsize=10, color=TEXT_MUTED)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=25, ha="right", fontsize=9)
    ax.grid(axis="y", color=BORDER, linewidth=0.8)
    ax.set_axisbelow(True)
    for s in ["top", "right"]:
        ax.spines[s].set_visible(False)
    ax.spines["left"].set_color(BORDER)
    ax.spines["bottom"].set_color(BORDER)
    ax.tick_params(length=0, labelsize=9, colors=TEXT_MUTED)
    ax.tick_params(axis="x", colors=TEXT_MAIN)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=3,
              frameon=False, fontsize=10)
    fig.subplots_adjust(left=0.06, right=0.99, top=0.98, bottom=0.28)
    return _mpl_fig_png(fig, plt)


def mpl_mini_donut(billable, nonbill, notworked):
    """Small per-QA donut with center utilization % and inside slice labels."""
    import math
    plt, _ = _mpl()
    vals = [max(0.0, billable), max(0.0, nonbill), max(0.0, notworked)]
    total = sum(vals) or 1.0
    util = billable / total * 100.0

    fig, ax = plt.subplots(figsize=(2.6, 2.6))
    wedges, _ = ax.pie(
        vals, colors=[BILLABLE_COLOR, NONBILL_COLOR, NOTWORKED_COLOR],
        startangle=90, counterclock=False,
        wedgeprops=dict(width=0.35, edgecolor="white", linewidth=2),
    )
    for w, v in zip(wedges, vals):
        ang = (w.theta2 + w.theta1) / 2.0
        pct = v / total * 100.0
        if pct >= 6:
            ax.text(0.82 * math.cos(math.radians(ang)),
                    0.82 * math.sin(math.radians(ang)),
                    f"{pct:.0f}%", ha="center", va="center", color="white",
                    fontsize=8, fontweight="bold")
    ax.text(0, 0.06, f"{util:.0f}%", ha="center", va="center",
            fontsize=17, fontweight="bold", color=TEXT_MAIN)
    ax.text(0, -0.20, "Utilization", ha="center", va="center",
            fontsize=7, color=TEXT_MUTED)
    ax.set_aspect("equal")
    fig.subplots_adjust(top=0.99, bottom=0.01, left=0.01, right=0.99)
    return _mpl_fig_png(fig, plt)


def build_export_chart_pngs(summary_df, billable, nonbill, notworked, chart_titles, qa_names):
    """Render every export chart with matplotlib and return
    (chart_pngs, mini_pngs) keyed exactly the way to_excel_bytes / to_pdf_bytes
    expect (chart_pngs by title, mini_pngs by QA name). Each render is guarded
    so one bad figure can't sink the whole export; a figure that fails is
    simply omitted and the builders fall back to their data tables for it."""
    chart_pngs, mini_pngs = {}, {}

    # Map each known chart title to its renderer. Titles come from
    # team_chart_figs so ordering/labels stay identical to the on-screen set.
    for title in chart_titles:
        try:
            low = title.lower()
            if "donut" in low or "utilization" in low:
                chart_pngs[title] = mpl_team_donut(billable, nonbill, notworked)
            elif "mix" in low:
                chart_pngs[title] = mpl_hours_mix(summary_df)
            else:  # "QA Comparison — Total Hours"
                chart_pngs[title] = mpl_bar_total(summary_df)
        except Exception:
            pass

    by_name = {r["QA Name"]: r for _, r in summary_df.iterrows()}
    for qa_name in qa_names:
        row = by_name.get(qa_name)
        if row is None:
            continue
        try:
            mini_pngs[qa_name] = mpl_mini_donut(
                row["Billable Hours"], row["Non-Billable Hours"], row["Hours Not Worked"])
        except Exception:
            pass

    return chart_pngs, mini_pngs


def to_excel_bytes(summary_df, detail_df, kpis, period_label, chart_titles, chart_pngs, qa_names, mini_pngs, include_images=True):
    """chart_pngs and mini_pngs are dicts of already-rendered PNG bytes,
    keyed by chart title / QA name respectively -- rendered once, shared with
    to_pdf_bytes, rather than re-rendered here from scratch. Pass an empty
    dict (not None) when include_images is False."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    wb = Workbook()

    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(start_color="0F1729", end_color="0F1729", fill_type="solid")
    title_font = Font(size=16, bold=True, color="0F1729")
    kpi_label_font = Font(size=10, bold=True, color="7A7F91")
    kpi_val_font = Font(size=18, bold=True, color="0F1729")
    note_font = Font(size=9, italic=True, color="C1543D")

    ws["B2"] = "QA Work Hours Dashboard"
    ws["B2"].font = title_font
    ws["B3"] = f"Period: {period_label}   |   Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws["B3"].font = Font(size=10, color="7A7F91")

    kpi_cells = [
        ("QA TEAM SIZE", f'{kpis["team_size"]}'),
        ("BILLABLE HOURS", f'{kpis["billable"]:.1f}'),
        ("NON-BILLABLE HOURS", f'{kpis["nonbill"]:.1f}'),
        ("UTILIZATION", f'{kpis["utilization"]:.1f}%'),
        ("TOTAL HOURS", f'{kpis["total"]:.1f}'),
    ]
    for i, (label, val) in enumerate(kpi_cells):
        col = get_column_letter(2 + i * 2)
        ws[f"{col}5"] = label
        ws[f"{col}5"].font = kpi_label_font
        ws[f"{col}6"] = val
        ws[f"{col}6"].font = kpi_val_font

    row_cursor = 9
    images_failed = False
    if include_images:
        for title in chart_titles:
            ws[f"B{row_cursor}"] = title
            ws[f"B{row_cursor}"].font = Font(size=12, bold=True, color="0F1729")
            row_cursor += 1
            png = chart_pngs.get(title)
            if png is not None:
                img = XLImage(io.BytesIO(png))
                # Size to each chart's natural aspect so the matplotlib images
                # aren't distorted: the team donut is square, the bar/mix
                # charts are wide.
                low = title.lower()
                if "donut" in low or "utilization" in low:
                    img.width, img.height = 340, 340
                else:
                    img.width, img.height = 700, 188
                ws.add_image(img, f"B{row_cursor}")
                row_cursor += 20
            else:
                images_failed = True
                ws[f"B{row_cursor}"] = "(chart image unavailable — see chart data in the tables below)"
                ws[f"B{row_cursor}"].font = note_font
                row_cursor += 2

        ws[f"B{row_cursor}"] = "Individual QA Breakdown"
        ws[f"B{row_cursor}"].font = Font(size=13, bold=True, color="0F1729")
        row_cursor += 2
        col_positions = ["B", "F", "J", "N"]
        start_row = row_cursor
        for idx, qa_name in enumerate(qa_names):
            col = col_positions[idx % 4]
            r = start_row + (idx // 4) * 14
            ws[f"{col}{r}"] = qa_name
            ws[f"{col}{r}"].font = Font(size=11, bold=True, color="0F1729")
            png = mini_pngs.get(qa_name)
            if png is not None:
                img = XLImage(io.BytesIO(png))
                img.width, img.height = 220, 220
                ws.add_image(img, f"{col}{r + 1}")
            else:
                images_failed = True
    else:
        ws[f"B{row_cursor}"] = "Chart images were skipped for this export. All figures are in the tables below."
        ws[f"B{row_cursor}"].font = note_font
        row_cursor += 2

    ws.column_dimensions["A"].width = 2

    ws_summary = wb.create_sheet("QA Summary")
    ws_summary.append(list(summary_df.columns))
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in summary_df.round(2).itertuples(index=False):
        ws_summary.append(list(row))
    for col_cells in ws_summary.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_summary.column_dimensions[col_cells[0].column_letter].width = max(12, length + 2)

    ws_detail = wb.create_sheet("Detail Data")
    ws_detail.append(list(detail_df.columns))
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in detail_df.itertuples(index=False):
        ws_detail.append(list(row))
    for col_cells in ws_detail.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_detail.column_dimensions[col_cells[0].column_letter].width = max(12, min(length + 2, 30))

    wb.save(buf)
    buf.seek(0)
    return buf, images_failed


DEFAULT_PDF_SECTIONS = {
    "Team Overview": True,
    "Individual QA Breakdown": True,
    "Per-QA Summary Table": True,
    # Daily Log is the one section that can genuinely run to hundreds of rows
    # and several pages on its own, so — unlike the other three — it starts
    # OFF by default. The person exporting opts in via the checkbox in the
    # Export panel rather than always getting a long log table by default.
    "Daily Log": False,
}


def to_pdf_bytes(summary_df, detail_df, kpis, period_label, chart_titles, chart_pngs, qa_names, mini_pngs,
                  include_images=True, sections=None):
    """chart_pngs and mini_pngs are the SAME already-rendered PNG dicts passed
    to to_excel_bytes -- rendering happens exactly once per figure, shared
    between both export formats, rather than once per format.

    `sections` controls which of the 4 top-level report sections
    ("Team Overview", "Individual QA Breakdown", "Per-QA Summary Table",
    "Daily Log") are actually written into the PDF -- this only affects the
    PDF. summary_df/detail_df themselves are unchanged, and to_excel_bytes is
    a completely separate function that always writes every section
    regardless of what's passed here, so the Excel export is unaffected by
    this toggle. A section that's turned off is skipped entirely (no heading,
    no placeholder, no reserved space) rather than merely hidden, so turning
    a section off also shortens the PDF. Defaults to DEFAULT_PDF_SECTIONS
    (everything on except Daily Log) when not provided, so any older caller
    that doesn't pass sections still gets the previous full-report behavior.

    --- Redesign note ---------------------------------------------------------
    This builder was fully re-styled into a "dashboard on paper" layout that
    mirrors the on-screen app: a dark navy header band, colored KPI tiles, a
    hero Team Utilization donut sitting beside supporting charts, an evenly
    spaced Individual QA donut grid that fills the page, and a polished summary
    table. Every page carries a footer rule + page number via the page
    decorator. NONE of the underlying data/logic changed -- the same
    summary_df / detail_df / kpis / chart_pngs / mini_pngs feed it, section
    toggles are honored identically, and the image-fallback tables are kept so
    a kaleido-less environment still produces a complete document.
    """
    if sections is None:
        sections = DEFAULT_PDF_SECTIONS
    show_team_overview = sections.get("Team Overview", True)
    show_individual_breakdown = sections.get("Individual QA Breakdown", True)
    show_summary_table = sections.get("Per-QA Summary Table", True)
    show_daily_log = sections.get("Daily Log", True)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, Image as RLImage, PageBreak, KeepTogether,
                                     Flowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    # ---- Palette (kept identical to the app's constants) --------------------
    C_INK = rl_colors.HexColor("#0F1729")
    C_INK2 = rl_colors.HexColor("#1B2540")
    C_PAGE = rl_colors.HexColor("#F4F5F8")
    C_BORDER = rl_colors.HexColor("#E7E9F0")
    C_TEXT = rl_colors.HexColor("#1F2333")
    C_MUTED = rl_colors.HexColor("#7A7F91")
    C_BILL = rl_colors.HexColor("#2F9E8F")
    C_NONBILL = rl_colors.HexColor("#E0A72E")
    C_NOTWORKED = rl_colors.HexColor("#C1543D")
    C_WHITE = rl_colors.white

    # Landscape A4 usable width with 12mm side margins ~= 273mm -> 774pt.
    PAGE_W, PAGE_H = landscape(A4)
    L_MARGIN = R_MARGIN = 12 * mm
    CONTENT_W = PAGE_W - L_MARGIN - R_MARGIN  # ~ 773pt

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=13 * mm, bottomMargin=13 * mm,
        leftMargin=L_MARGIN, rightMargin=R_MARGIN,
        title="QA Work Hours Dashboard Report",
    )

    styles = getSampleStyleSheet()

    def _p(name, **kw):
        base = kw.pop("parent", styles["Normal"])
        return ParagraphStyle(name, parent=base, **kw)

    st_section = _p("SecHead", fontSize=15, leading=18, textColor=C_INK,
                    fontName="Helvetica-Bold", spaceBefore=0, spaceAfter=2)
    st_section_sub = _p("SecSub", fontSize=8.5, leading=11, textColor=C_MUTED,
                        fontName="Helvetica", spaceAfter=4)
    st_chart_cap = _p("ChartCap", fontSize=9.5, leading=12, textColor=C_INK,
                      fontName="Helvetica-Bold", alignment=TA_CENTER)
    st_qa_name = _p("QAName", fontSize=9.5, leading=12, textColor=C_INK,
                    fontName="Helvetica-Bold", alignment=TA_CENTER)
    st_note = _p("Note", fontSize=8.5, leading=11, textColor=C_MUTED,
                 fontName="Helvetica-Oblique")

    # ------------------------------------------------------------------ #
    #  Page furniture: header band (first page) + footer on every page   #
    # ------------------------------------------------------------------ #
    gen_str = datetime.now().strftime('%d %b %Y, %H:%M')

    def _draw_footer(canvas, doc_):
        canvas.saveState()
        y = 9 * mm
        canvas.setStrokeColor(C_BORDER)
        canvas.setLineWidth(0.6)
        canvas.line(L_MARGIN, y + 4, PAGE_W - R_MARGIN, y + 4)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(C_MUTED)
        canvas.drawString(L_MARGIN, y - 4, "QA Work Hours Dashboard")
        canvas.drawCentredString(PAGE_W / 2.0, y - 4, f"Period: {period_label}")
        canvas.drawRightString(PAGE_W - R_MARGIN, y - 4, f"Page {doc_.page}")
        canvas.restoreState()

    def _first_page(canvas, doc_):
        # Full-bleed navy header banner across the very top of page 1.
        canvas.saveState()
        band_h = 20 * mm
        top = PAGE_H
        canvas.setFillColor(C_INK)
        canvas.rect(0, top - band_h, PAGE_W, band_h, stroke=0, fill=1)
        # thin teal accent line under the band
        canvas.setFillColor(C_BILL)
        canvas.rect(0, top - band_h - 2, PAGE_W, 2, stroke=0, fill=1)
        # eyebrow + title
        canvas.setFillColor(rl_colors.HexColor("#8B93AD"))
        canvas.setFont("Helvetica-Bold", 8)
        canvas.drawString(L_MARGIN, top - 8 * mm, "QA MANAGEMENT SYSTEM  ·  ANALYTICS REPORT")
        canvas.setFillColor(C_WHITE)
        canvas.setFont("Helvetica-Bold", 19)
        canvas.drawString(L_MARGIN, top - 15 * mm, "QA Work Hours Dashboard")
        # right-aligned meta
        canvas.setFillColor(rl_colors.HexColor("#B9C0D4"))
        canvas.setFont("Helvetica", 8.5)
        canvas.drawRightString(PAGE_W - R_MARGIN, top - 8.5 * mm, f"Period: {period_label}")
        canvas.drawRightString(PAGE_W - R_MARGIN, top - 13.5 * mm, f"Generated: {gen_str}")
        canvas.restoreState()
        _draw_footer(canvas, doc_)

    # ------------------------------------------------------------------ #
    #  Small reusable flowables                                          #
    # ------------------------------------------------------------------ #
    class HRule(Flowable):
        """A thin full-width divider rule."""
        def __init__(self, width, color=C_BORDER, thickness=0.7):
            super().__init__()
            self.width = width
            self.color = color
            self.thickness = thickness

        def wrap(self, aw, ah):
            return (self.width, self.thickness + 2)

        def draw(self):
            self.canv.setStrokeColor(self.color)
            self.canv.setLineWidth(self.thickness)
            self.canv.line(0, 1, self.width, 1)

    class BarMeter(Flowable):
        """A slim rounded progress bar: a light track with a colored fill
        showing `frac` (0..1). Used in the Utilization Breakdown panel."""
        def __init__(self, width, frac, color, height=7, track=rl_colors.HexColor("#EEF0F4")):
            super().__init__()
            self.width = width
            self.frac = max(0.0, min(1.0, frac))
            self.color = color
            self.height = height
            self.track = track

        def wrap(self, aw, ah):
            return (self.width, self.height)

        def draw(self):
            r = self.height / 2.0
            c = self.canv
            c.setFillColor(self.track)
            c.roundRect(0, 0, self.width, self.height, r, stroke=0, fill=1)
            fill_w = max(self.height, self.width * self.frac) if self.frac > 0 else 0
            if fill_w > 0:
                c.setFillColor(self.color)
                c.roundRect(0, 0, fill_w, self.height, r, stroke=0, fill=1)

    def _kpi_tile(label, value, sub, accent):
        """One rounded KPI tile as a single-cell Table so it gets a border,
        a colored accent strip up top, and stacked label/value/sub text."""
        inner = Table(
            [[Paragraph(label, _p("kl", fontSize=7.5, leading=9, textColor=C_MUTED,
                                  fontName="Helvetica-Bold"))],
             [Paragraph(value, _p("kv", fontSize=17, leading=19, textColor=C_INK,
                                  fontName="Helvetica-Bold"))],
             [Paragraph(sub, _p("ks", fontSize=7, leading=9, textColor=C_MUTED,
                                fontName="Helvetica"))]],
            colWidths=[CONTENT_W / 6.0 - 6],
        )
        inner.setStyle(TableStyle([
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (0, 0), 3),
            ("BOTTOMPADDING", (0, 1), (0, 1), 2),
            ("BOTTOMPADDING", (0, 2), (0, 2), 0),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("LINEABOVE", (0, 0), (0, 0), 2.4, accent),
            ("TOPPADDING", (0, 0), (0, 0), 4),
        ]))
        return inner

    # ------------------------------------------------------------------ #
    #  Build the story                                                   #
    # ------------------------------------------------------------------ #
    elements = []
    images_failed = False
    team_overview_rendered = False
    individual_breakdown_rendered = False
    summary_table_rendered = False

    # Leave room below the printed header band on page 1.
    elements.append(Spacer(1, 15 * mm))

    # ---- KPI tile strip (always shown) ----
    tile_specs = [
        ("QA TEAM SIZE", f'{kpis["team_size"]}', "active members", C_INK),
        ("TOTAL HOURS", f'{kpis["total"]:,.1f}', "hrs this period", C_INK2),
        ("BILLABLE HOURS", f'{kpis["billable"]:,.1f}', "hrs logged", C_BILL),
        ("NON-BILLABLE", f'{kpis["nonbill"]:,.1f}', "hrs logged", C_NONBILL),
        ("UTILIZATION", f'{kpis["utilization"]:.1f}%', "billable share", C_NOTWORKED),
    ]
    # 5 real tiles + 1 spacer keeps them the same width as the 6-col grid used
    # elsewhere; instead just spread 5 evenly across the full width.
    tile_w = CONTENT_W / 5.0
    tile_cells = [_kpi_tile(*spec) for spec in tile_specs]
    kpi_strip = Table([tile_cells], colWidths=[tile_w] * 5)
    kpi_strip.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, -1), C_WHITE),
        ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.8, C_BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
    ]))
    elements.append(kpi_strip)
    elements.append(Spacer(1, 12))

    # ================================================================== #
    #  SECTION 1 — TEAM OVERVIEW                                          #
    # ================================================================== #
    if show_team_overview:
        # Collect the rendered chart PNGs by title up front.
        chart_by_title = {}
        if include_images:
            for title in chart_titles:
                png = chart_pngs.get(title)
                if png is not None:
                    chart_by_title[title] = png
                else:
                    images_failed = True

        donut_title = next((t for t in chart_titles if "Donut" in t or "Utilization" in t), None)
        bar_title = next((t for t in chart_titles if "Comparison" in t or "Total Hours" in t), None)
        mix_title = next((t for t in chart_titles if "Mix" in t), None)

        # ---- Derived figures for the insight panel -----------------------
        bill = float(kpis.get("billable", 0) or 0)
        non = float(kpis.get("nonbill", 0) or 0)
        tot = float(kpis.get("total", 0) or 0)
        notw = max(0.0, tot - bill - non)
        denom = tot if tot > 0 else 1.0
        bill_pct, non_pct, notw_pct = bill / denom * 100, non / denom * 100, notw / denom * 100

        sorted_summary = summary_df.sort_values("Total Hours", ascending=False)
        top_name = sorted_summary.iloc[0]["QA Name"] if len(sorted_summary) else "\u2014"
        top_hours = float(sorted_summary.iloc[0]["Total Hours"]) if len(sorted_summary) else 0.0
        team_n = int(kpis.get("team_size", len(summary_df)) or len(summary_df))
        avg_per_qa = (tot / team_n) if team_n else 0.0

        # =============================================================== #
        #  PAGE 1 — Team utilization donut + breakdown insight panel      #
        # =============================================================== #
        elements.append(Paragraph("Team Overview", st_section))
        elements.append(Paragraph(
            "How the team's logged hours split across billable, non-billable, and not-worked.",
            st_section_sub))
        elements.append(HRule(CONTENT_W))
        elements.append(Spacer(1, 14))

        # ---- Left: donut card ----
        if donut_title and donut_title in chart_by_title:
            donut_body = RLImage(io.BytesIO(chart_by_title[donut_title]), width=286, height=250)
        else:
            donut_body = Paragraph(
                f"<b>{kpis.get('utilization', 0):.1f}%</b> utilization",
                _p("dnu", fontSize=13, leading=16, textColor=C_INK, alignment=TA_CENTER))
        left_card = Table(
            [[Paragraph("Team Utilization Split", st_chart_cap)], [donut_body]],
            colWidths=[338])
        left_card.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, -1), C_WHITE),
            ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
            ("ROUNDEDCORNERS", [8, 8, 8, 8]),
            ("TOPPADDING", (0, 0), (0, 0), 12),
            ("BOTTOMPADDING", (0, 0), (0, 0), 4),
            ("TOPPADDING", (0, 1), (0, 1), 2),
            ("BOTTOMPADDING", (0, 1), (0, 1), 14),
        ]))

        # ---- Right: "Utilization Breakdown" insight panel ----
        meter_w = 360

        def _breakdown_row(color, label, hours, pct):
            head = Table(
                [[Paragraph(f'<font color="{color}">&#9632;</font> &nbsp;<b>{label}</b>',
                            _p("brl", fontSize=10.5, leading=13, textColor=C_TEXT)),
                  Paragraph(f'<b>{hours:,.1f}</b> hrs &nbsp;&middot;&nbsp; {pct:.1f}%',
                            _p("brv", fontSize=10.5, leading=13, textColor=C_MUTED, alignment=TA_RIGHT))]],
                colWidths=[meter_w * 0.5, meter_w * 0.5])
            head.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            return [head, Spacer(1, 2),
                    BarMeter(meter_w, pct / 100.0, rl_colors.HexColor(color)),
                    Spacer(1, 12)]

        panel_flow = [Paragraph("Utilization Breakdown", _p(
            "bdh", fontSize=12, leading=15, textColor=C_INK, fontName="Helvetica-Bold"))]
        panel_flow.append(Spacer(1, 12))
        panel_flow += _breakdown_row("#2F9E8F", "Billable", bill, bill_pct)
        panel_flow += _breakdown_row("#E0A72E", "Non-Billable", non, non_pct)
        panel_flow += _breakdown_row("#C1543D", "Not Worked", notw, notw_pct)
        panel_flow.append(HRule(meter_w))
        panel_flow.append(Spacer(1, 10))

        highlight = Table([
            [Paragraph("HIGHEST TOTAL", _p("hk", fontSize=7.5, leading=10, textColor=C_MUTED,
                                          fontName="Helvetica-Bold")),
             Paragraph("AVG PER QA", _p("hk2", fontSize=7.5, leading=10, textColor=C_MUTED,
                                        fontName="Helvetica-Bold"))],
            [Paragraph(f"<b>{top_name}</b>", _p("hv", fontSize=11, leading=14, textColor=C_INK)),
             Paragraph(f"<b>{avg_per_qa:,.1f}</b> hrs", _p("hv2", fontSize=11, leading=14, textColor=C_INK))],
            [Paragraph(f"{top_hours:,.1f} hrs", _p("hs", fontSize=8.5, leading=11, textColor=C_MUTED)),
             Paragraph(f"across {team_n} QAs", _p("hs2", fontSize=8.5, leading=11, textColor=C_MUTED))],
        ], colWidths=[meter_w * 0.5, meter_w * 0.5])
        highlight.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 1), ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        panel_flow.append(highlight)

        right_card = Table([[panel_flow]], colWidths=[398])
        right_card.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, -1), C_WHITE),
            ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
            ("ROUNDEDCORNERS", [8, 8, 8, 8]),
            ("LEFTPADDING", (0, 0), (-1, -1), 20),
            ("RIGHTPADDING", (0, 0), (-1, -1), 20),
            ("TOPPADDING", (0, 0), (-1, -1), 16),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
        ]))

        page1_row = Table([[left_card, right_card]], colWidths=[346, 406])
        page1_row.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 8),
            ("RIGHTPADDING", (1, 0), (1, 0), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(page1_row)

        # =============================================================== #
        #  PAGE 2 — QA comparison + hours mix, each a full-width card      #
        # =============================================================== #
        have_team_charts = (bar_title in chart_by_title) or (mix_title in chart_by_title)
        if have_team_charts:
            elements.append(PageBreak())
            elements.append(Spacer(1, 2))
            elements.append(Paragraph("Team Charts", st_section))
            elements.append(Paragraph(
                "Total hours per QA, and each person's billable / non-billable / not-worked mix.",
                st_section_sub))
            elements.append(HRule(CONTENT_W))
            elements.append(Spacer(1, 12))

            def _wide_chart_card(title_text, png, img_h):
                body = RLImage(io.BytesIO(png), width=CONTENT_W - 44, height=img_h)
                card = Table([[Paragraph(title_text, st_chart_cap)], [body]],
                             colWidths=[CONTENT_W])
                card.setStyle(TableStyle([
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BACKGROUND", (0, 0), (-1, -1), C_WHITE),
                    ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
                    ("ROUNDEDCORNERS", [8, 8, 8, 8]),
                    ("TOPPADDING", (0, 0), (0, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (0, 0), 3),
                    ("TOPPADDING", (0, 1), (0, 1), 2),
                    ("BOTTOMPADDING", (0, 1), (0, 1), 10),
                ]))
                # KeepTogether so a card's title never separates from its chart
                # across a page break (that was leaving an orphaned heading).
                return KeepTogether([card])

            if bar_title in chart_by_title:
                elements.append(_wide_chart_card("QA Comparison — Total Hours",
                                                 chart_by_title[bar_title], 167))
                elements.append(Spacer(1, 12))
            if mix_title in chart_by_title:
                elements.append(_wide_chart_card("Hours Mix — Billable / Non-Billable / Not Worked",
                                                 chart_by_title[mix_title], 176))

        # ---- Fallback when NO chart images are available at all ----------
        if not chart_by_title:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                "Chart images could not be rendered for this export. The figures "
                "below are the same data the on-screen charts show, as tables.",
                st_note))
            elements.append(Spacer(1, 8))
            comp_df = summary_df[["QA Name", "Total Hours"]].sort_values("Total Hours", ascending=False)
            mix_df = summary_df[["QA Name", "Billable Hours", "Non-Billable Hours", "Hours Not Worked"]]
            comp_data = [["QA Name", "Total Hours"]] + comp_df.round(1).astype(str).values.tolist()
            comp_tbl = Table(comp_data, colWidths=[240, 130])
            mix_data = [["QA Name", "Billable", "Non-Billable", "Not Worked"]] + mix_df.round(1).astype(str).values.tolist()
            mix_tbl = Table(mix_data, colWidths=[150, 80, 90, 80])
            for tbl in (comp_tbl, mix_tbl):
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), C_INK),
                    ("TEXTCOLOR", (0, 0), (-1, 0), C_WHITE),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.4, C_BORDER),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_PAGE]),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]))
            side = Table([[
                Table([[Paragraph("QA Comparison — Total Hours", st_chart_cap)], [comp_tbl]]),
                Table([[Paragraph("Hours Mix", st_chart_cap)], [mix_tbl]]),
            ]], colWidths=[380, 400])
            side.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            elements.append(side)

        team_overview_rendered = True

    # ================================================================== #
    #  SECTION 2 — INDIVIDUAL QA BREAKDOWN                                #
    # ================================================================== #
    if show_individual_breakdown:
        if team_overview_rendered:
            elements.append(PageBreak())
            elements.append(Spacer(1, 2))
        elements.append(Paragraph("Individual QA Breakdown", st_section))
        elements.append(Paragraph(
            "Center figure = each QA's Billable Hours ÷ Total Hours (their own billable share) for the period.",
            st_section_sub))
        elements.append(HRule(CONTENT_W))
        elements.append(Spacer(1, 10))

        # Adaptive grid so the donuts fill the page regardless of team size.
        n = len(qa_names)
        per_row = 4 if n > 3 else max(1, n)
        if n <= 4:
            img_size = 168
        elif n <= 8:
            img_size = 150
        else:
            img_size = 128
        col_w = CONTENT_W / per_row

        # Build one "card" flowable per QA (name + donut, boxed).
        def _qa_card(qa_name):
            img = None
            if include_images:
                png = mini_pngs.get(qa_name)
                if png is not None:
                    img = RLImage(io.BytesIO(png), width=img_size, height=img_size)
            body = img if img is not None else Paragraph(
                "(image<br/>unavailable)", _p("nu", fontSize=8, leading=10,
                                              textColor=C_MUTED, alignment=TA_CENTER))
            card = Table([[Paragraph(qa_name, st_qa_name)], [body]],
                         colWidths=[col_w - 10])
            card.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (-1, -1), C_WHITE),
                ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
                ("ROUNDEDCORNERS", [6, 6, 6, 6]),
                ("TOPPADDING", (0, 0), (0, 0), 7),
                ("BOTTOMPADDING", (0, 0), (0, 0), 2),
                ("TOPPADDING", (0, 1), (0, 1), 0),
                ("BOTTOMPADDING", (0, 1), (0, 1), 7),
            ]))
            return card

        any_mini_image = include_images and any(mini_pngs.get(q) is not None for q in qa_names)
        if include_images and not any_mini_image:
            images_failed = True

        # Assemble rows of cards.
        cards = [_qa_card(q) for q in qa_names]
        row = []
        grid_rows = []
        for c in cards:
            row.append(c)
            if len(row) == per_row:
                grid_rows.append(row)
                row = []
        if row:
            while len(row) < per_row:  # pad the last row so widths stay even
                row.append(Paragraph("", styles["Normal"]))
            grid_rows.append(row)

        if grid_rows:
            grid = Table(grid_rows, colWidths=[col_w] * per_row)
            grid.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            elements.append(grid)

        if not any_mini_image:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(
                "Donut images could not be rendered in this environment. Per-QA "
                "utilization (the number the donuts show) is in the table below.",
                st_note))
            elements.append(Spacer(1, 6))
            util_df = summary_df[["QA Name", "Billable Hours", "Non-Billable Hours",
                                   "Hours Not Worked", "Total Hours", "Utilization %"]]
            util_data = [["QA Name", "Billable", "Non-Billable", "Not Worked", "Total", "Utilization %"]]
            for r in util_df.round(1).values.tolist():
                util_data.append([r[0], f"{r[1]:.1f}", f"{r[2]:.1f}", f"{r[3]:.1f}", f"{r[4]:.1f}", f"{r[5]:.1f}%"])
            n_c = 6
            util_tbl = Table(util_data, colWidths=[CONTENT_W / n_c] * n_c, repeatRows=1)
            util_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_INK),
                ("TEXTCOLOR", (0, 0), (-1, 0), C_WHITE),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, C_BORDER),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_PAGE]),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            elements.append(util_tbl)

        individual_breakdown_rendered = True

    # ================================================================== #
    #  SECTION 3 — PER-QA SUMMARY TABLE                                   #
    # ================================================================== #
    if show_summary_table:
        if team_overview_rendered or individual_breakdown_rendered:
            elements.append(PageBreak())
            elements.append(Spacer(1, 2))
        elements.append(Paragraph("Per-QA Summary Table", st_section))
        elements.append(Paragraph(
            "Aggregated hours and utilization per QA for the selected period.",
            st_section_sub))
        elements.append(HRule(CONTENT_W))
        elements.append(Spacer(1, 10))

        cols = list(summary_df.columns)
        n_cols = len(cols)
        header = [Paragraph(f"<b>{c}</b>", _p("th", fontSize=10, leading=12,
                            textColor=C_WHITE, alignment=TA_CENTER)) for c in cols]
        body_rows = summary_df.round(1).astype(str).values.tolist()
        table_data = [header] + body_rows

        col_w = CONTENT_W / n_cols
        per_qa_tbl = Table(table_data, colWidths=[col_w] * n_cols, repeatRows=1)

        # Taller, generously padded rows so the table reads as a designed
        # report table filling the page rather than a small dense grid.
        n_body = len(body_rows)
        row_pad = 10 if n_body <= 10 else (7 if n_body <= 18 else 5)
        per_qa_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), C_INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), C_WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 1), (-1, -1), C_TEXT),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
            ("LINEBEFORE", (0, 0), (-1, -1), 0.5, C_BORDER),
            ("LINEAFTER", (-1, 0), (-1, -1), 0.5, C_BORDER),
            ("LINEABOVE", (0, 0), (-1, 0), 0.5, C_INK),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, rl_colors.HexColor("#F7F8FB")]),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), row_pad),
            ("BOTTOMPADDING", (0, 0), (-1, -1), row_pad),
            ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
        ]))
        elements.append(per_qa_tbl)
        summary_table_rendered = True

    # ================================================================== #
    #  SECTION 4 — DAILY LOG (unchanged behavior, restyled header)       #
    # ================================================================== #
    if show_daily_log:
        if team_overview_rendered or individual_breakdown_rendered or summary_table_rendered:
            elements.append(PageBreak())
            elements.append(Spacer(1, 2))
        elements.append(Paragraph("Daily Log", st_section))
        elements.append(Paragraph(
            f"{len(detail_df):,} rows · sorted by date, most recent first.",
            st_section_sub))
        elements.append(HRule(CONTENT_W))
        elements.append(Spacer(1, 8))

        log_for_pdf = detail_df.copy().sort_values("Date", ascending=False)
        log_for_pdf["Date"] = pd.to_datetime(log_for_pdf["Date"]).dt.strftime("%Y-%m-%d")
        log_cols = ["Date", "Day", "QA Name", "Billable Hours", "Non-Billable Hours",
                    "Hours Not Worked", "Total Hours", "Comment"]
        log_for_pdf = log_for_pdf[log_cols]

        MAX_PDF_LOG_ROWS = 500
        truncated = len(log_for_pdf) > MAX_PDF_LOG_ROWS
        log_for_pdf_show = log_for_pdf.head(MAX_PDF_LOG_ROWS)

        log_table_data = [log_cols] + log_for_pdf_show.round(2).astype(str).values.tolist()
        log_col_widths = [64, 38, 92, 66, 78, 72, 66, 297]
        log_tbl = Table(log_table_data, colWidths=log_col_widths, repeatRows=1)
        log_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), C_INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), C_WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.4, C_BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_PAGE]),
            ("ALIGN", (0, 0), (6, -1), "CENTER"),
            ("ALIGN", (7, 0), (7, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(log_tbl)
        if truncated:
            elements.append(Spacer(1, 6))
            elements.append(Paragraph(
                f"Showing the most recent {MAX_PDF_LOG_ROWS:,} of {len(log_for_pdf):,} rows. "
                "Download the Excel export for the complete daily log.",
                st_note))

    if not any([show_team_overview, show_individual_breakdown, show_summary_table, show_daily_log]):
        elements.append(Paragraph(
            "No report sections were selected for this export. Check at least one of "
            "Team Overview, Individual QA Breakdown, Per-QA Summary Table, or Daily Log "
            "in the Export panel and click Prepare Exports again.",
            st_note))

    # First page gets the printed header band; every page gets the footer.
    doc.build(elements, onFirstPage=_first_page, onLaterPages=_draw_footer)
    buf.seek(0)
    return buf, images_failed


# ============================================================================
# APP HEADER / UPLOAD
# ============================================================================

st.markdown("""
<div class="top-banner">
    <div class="eyebrow">QA MANAGEMENT SYSTEM &middot; V2</div>
    <div class="title">QA Hours Dashboard</div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader("Upload QA Work Hours Excel file (.xlsx)", type=["xlsx"])

if uploaded is None:
    st.info("\U0001F446 Upload your QA hours tracker workbook to begin. Each QA's own sheet is detected and combined automatically. Only data up to today's date will ever be shown.")
    st.stop()

with st.spinner("Reading and analyzing the workbook..."):
    data = parse_workbook(uploaded)

if data.empty:
    st.error("Couldn't find any recognizable QA hour records in this workbook (or all rows were in the future). Please check the sheet format.")
    st.stop()

# ============================================================================
# SIDEBAR FILTERS
# ============================================================================
st.sidebar.header("\U0001F50D Filters")

# "Clear All" (button lives at the very bottom of the sidebar, see below)
# has to take effect here, BEFORE any filter widget below is instantiated —
# Streamlit widgets read their bound session_state key at creation time, so
# clearing state after the widgets already rendered this run would only be
# visible on the NEXT interaction, not immediately. The button itself sets
# this flag and calls st.rerun(); this block is what actually acts on it.
if st.session_state.get("_clear_all_filters"):
    st.session_state["_clear_all_filters"] = False
    st.session_state["qa_multiselect"] = []
    # Empty each, matching exactly what that filter's own individual Clear
    # button does — Clear All is just "press every Clear button at once".
    for _k in ("years_multiselect", "months_multiselect", "weeks_multiselect"):
        st.session_state[_k] = []

view_mode = st.sidebar.radio("View by", ["Daily", "Weekly", "Monthly", "Yearly"], index=2)

# ---- QA selection: a real closed dropdown (multiselect), with quick All/None ----
all_qas = sorted(data["QA Name"].unique().tolist())

if "qa_multiselect" not in st.session_state:
    st.session_state.qa_multiselect = all_qas.copy()
# keep state in sync if a new file introduces different QAs
st.session_state.qa_multiselect = [qa for qa in st.session_state.qa_multiselect if qa in all_qas]

if st.sidebar.button("Clear QAs", use_container_width=True):
    st.session_state.qa_multiselect = []

sel_qas = st.sidebar.multiselect("QA Team Members", all_qas, key="qa_multiselect")

st.sidebar.markdown("---")

# ---- Time period selection depending on view mode ----
years_available = sorted(data["Year"].dropna().unique().astype(int).tolist())

def _sidebar_multiselect_with_clear(label, options, state_key, clear_label, format_func=None):
    """A multiselect that defaults to "all selected" the first time it's seen,
    persists its selection in session_state under state_key (same pattern as
    the QA Team Members widget above), and renders its own small "Clear <x>"
    button directly above it. Matches the existing, working Clear QAs
    pattern exactly: the button mutates st.session_state[state_key] BEFORE
    the multiselect widget below it is instantiated in this same script run,
    so the widget picks up the cleared value immediately with no st.rerun()
    needed — unlike passing default= each run, which would snap back to
    "all selected" since nothing would ever persist a cleared state.
    """
    if state_key not in st.session_state:
        st.session_state[state_key] = list(options)
    # keep state in sync if the available options changed (e.g. new file, or
    # Years selection narrowing which Months/Weeks exist)
    st.session_state[state_key] = [v for v in st.session_state[state_key] if v in options]

    if st.sidebar.button(clear_label, use_container_width=True, key=f"{state_key}_clear_btn"):
        st.session_state[state_key] = []

    kwargs = {"key": state_key}
    if format_func is not None:
        kwargs["format_func"] = format_func
    return st.sidebar.multiselect(label, options, **kwargs)


if view_mode == "Yearly":
    sel_years = _sidebar_multiselect_with_clear("Years", years_available, "years_multiselect", "Clear Years")
    df_period = data[data["Year"].isin(sel_years)]
    period_label = ", ".join(str(y) for y in sel_years) if sel_years else "No years selected"

elif view_mode == "Monthly":
    sel_years = _sidebar_multiselect_with_clear("Years", years_available, "years_multiselect", "Clear Years")
    months_in_scope = [m for m in MONTH_ORDER if m in data[data["Year"].isin(sel_years)]["Month"].unique()]
    sel_months = _sidebar_multiselect_with_clear("Months", months_in_scope, "months_multiselect", "Clear Months")
    df_period = data[data["Year"].isin(sel_years) & data["Month"].isin(sel_months)]
    # Dynamic label: "Feb 2026" for a single month/year, "Feb, Mar (2026)"
    # for several months in one year, "Feb 2025, Mar 2026" if years differ —
    # never a fixed placeholder string.
    if not sel_months or not sel_years:
        period_label = "No months selected"
    elif len(sel_years) == 1:
        year_txt = str(sel_years[0])
        if len(sel_months) == 1:
            period_label = f"{sel_months[0]} {year_txt}"
        else:
            period_label = f"{', '.join(sel_months)} ({year_txt})"
    else:
        period_label = f"{', '.join(sel_months)} ({', '.join(str(y) for y in sel_years)})"

elif view_mode == "Weekly":
    sel_years = _sidebar_multiselect_with_clear("Years", years_available, "years_multiselect", "Clear Years")
    df_y = data[data["Year"].isin(sel_years)]
    df_y = df_y.assign(_week=df_y["Date"].dt.to_period("W"))
    week_options = sorted(df_y["_week"].unique())
    week_labels = {w: f"{w.start_time.strftime('%d %b')} – {w.end_time.strftime('%d %b %Y')}" for w in week_options}
    sel_weeks = _sidebar_multiselect_with_clear(
        "Weeks", week_options, "weeks_multiselect", "Clear Weeks",
        format_func=lambda w: week_labels.get(w, str(w)),
    )
    df_period = df_y[df_y["_week"].isin(sel_weeks)].drop(columns="_week")
    if not sel_weeks:
        period_label = "No weeks selected"
    elif len(sel_weeks) == 1:
        period_label = week_labels.get(sel_weeks[0], str(sel_weeks[0]))
    else:
        period_label = f"{len(sel_weeks)} weeks selected"

else:  # Daily -> calendar-style multi-date picker
    min_d, max_d = data["Date"].min().date(), min(data["Date"].max().date(), TODAY.date())
    sel_dates = st.sidebar.date_input(
        "Select date(s)", value=(min_d, max_d), min_value=min_d, max_value=max_d,
    )
    if isinstance(sel_dates, (tuple, list)) and len(sel_dates) == 2:
        d_start, d_end = sel_dates
        df_period = data[(data["Date"].dt.date >= d_start) & (data["Date"].dt.date <= d_end)]
        if d_start == d_end:
            period_label = d_start.strftime("%d %b %Y")
        else:
            period_label = f"{d_start.strftime('%d %b %Y')} – {d_end.strftime('%d %b %Y')}"
    elif isinstance(sel_dates, (tuple, list)) and len(sel_dates) == 1:
        d_only = sel_dates[0]
        df_period = data[data["Date"].dt.date == d_only]
        period_label = d_only.strftime("%d %b %Y")
    else:
        df_period = data[data["Date"].dt.date == sel_dates]
        period_label = sel_dates.strftime("%d %b %Y")

st.sidebar.markdown("---")
if st.sidebar.button("\U0001F9F9 Clear All Filters", use_container_width=True):
    st.session_state["_clear_all_filters"] = True
    st.rerun()

df_period = df_period[df_period["QA Name"].isin(sel_qas)]

if df_period.empty:
    st.warning("No data for the selected filters. Adjust the QA / time filters in the sidebar.")
    st.stop()

qa_colors = qa_color_map(all_qas)

# ============================================================================
# KPI SUMMARY
# ============================================================================

# "QA Team Size" reflects how many QAs are CHECKED in the filter — not just
# how many happen to have logged rows within the current date/month/year
# window (a QA with zero hours in a narrow window should still count as
# selected, matching what the sidebar shows).
team_size = len(sel_qas)
billable_total = df_period["Billable Hours"].sum()
nonbill_total = df_period["Non-Billable Hours"].sum()
notworked_total = df_period["Hours Not Worked"].sum()
total_hours = df_period["Total Hours"].sum()
utilization = (billable_total / total_hours * 100) if total_hours > 0 else 0
days_logged = df_period["Date"].nunique()
avg_hours_per_day = (total_hours / (df_period.groupby("QA Name")["Date"].nunique().sum())) if len(df_period) else 0

kpis = dict(team_size=team_size, billable=billable_total, nonbill=nonbill_total,
            utilization=utilization, total=total_hours)

st.markdown(f'<div class="panel-title" style="margin-bottom:12px;">QA &mdash; {period_label}</div>', unsafe_allow_html=True)

k1, k2, k3, k4, k5, k6 = st.columns(6)
kpi_cells = [
    (k1, "QA TEAM SIZE", f"{team_size}", "active members"),
    (k2, "TOTAL HOURS", f"{total_hours:,.1f}", f"across {days_logged} days"),
    (k3, "BILLABLE SHARE", f"{utilization:.1f}%", f"{billable_total:,.0f} of {total_hours:,.0f} hrs"),
    (k4, "AVG HOURS / DAY", f"{avg_hours_per_day:.2f}", "per active QA-day"),
    (k5, "NON-BILLABLE HOURS", f"{nonbill_total:,.1f}", "hrs logged"),
    (k6, "HOURS NOT WORKED", f"{notworked_total:,.1f}", "hrs logged"),
]
for col, label, val, sub in kpi_cells:
    with col:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{val}</div>
            <div class="kpi-sub">{sub}</div>
        </div>
        """, unsafe_allow_html=True)

st.markdown('<div class="kpi-row-spacer"></div>', unsafe_allow_html=True)

# ============================================================================
# CHARTS - TEAM LEVEL
# ============================================================================

fig_donut = donut_chart(billable_total, nonbill_total, notworked_total)
fig_bar = bar_chart_by_qa(df_period)
fig_mix = hours_mix_chart(df_period)

with st.container(border=True):
    st.plotly_chart(fig_bar, use_container_width=True, config=PLOTLY_CONFIG)

with st.container(border=True):
    st.plotly_chart(fig_mix, use_container_width=True, config=PLOTLY_CONFIG)

# Donut, centered
d1, d2, d3 = st.columns([1, 2, 1])
with d2:
    with st.container(border=True):
        st.plotly_chart(fig_donut, use_container_width=True, config=PLOTLY_CONFIG)

team_chart_figs = [
    ("QA Comparison — Total Hours", fig_bar),
    ("Hours Mix", fig_mix),
    ("Team Utilization Split (Donut)", fig_donut),
]

# ============================================================================
# PER-QA CARDS
# ============================================================================
st.markdown(
    '<div class="panel-title">Individual QA Breakdown</div>'
    '<div class="panel-sub" style="margin-top:2px;">'
    'The number in the center of each donut is that QA\u2019s <b>Billable Hours \u00f7 Total Hours</b> '
    '(billable share of their own hours) for the period selected above \u2014 not a share of the whole team.'
    '</div>',
    unsafe_allow_html=True,
)
st.write("")

qa_summary = df_period.groupby("QA Name").agg(
    **{
        "Billable Hours": ("Billable Hours", "sum"),
        "Non-Billable Hours": ("Non-Billable Hours", "sum"),
        "Hours Not Worked": ("Hours Not Worked", "sum"),
        "Total Hours": ("Total Hours", "sum"),
        "Days Logged": ("Date", "nunique"),
    }
).reset_index()
qa_summary["Utilization %"] = np.where(
    qa_summary["Total Hours"] > 0,
    qa_summary["Billable Hours"] / qa_summary["Total Hours"] * 100,
    0,
)
# Canonical sort applied everywhere in the app: Total Hours, descending.
# qa_summary feeds the breakdown cards below, the "View Per-QA Summary
# Table" expander, and both the Excel and PDF exports — sorting it here
# once keeps all of those consistent with the bar chart and stacked chart.
qa_summary = qa_summary.sort_values("Total Hours", ascending=False)

n_cols = 3
qa_rows = [qa_summary.iloc[i:i + n_cols] for i in range(0, len(qa_summary), n_cols)]

qa_mini_figs = []
for chunk in qa_rows:
    cols = st.columns(n_cols)
    for col, (_, row) in zip(cols, chunk.iterrows()):
        with col:
            st.markdown(f"""
            <div class="qa-card">
                <div class="qa-name">{row['QA Name']}</div>
                <div class="kpi-sub">Days logged: {int(row['Days Logged'])} &nbsp;|&nbsp; Total: {row['Total Hours']:.1f} hrs</div>
            </div>
            """, unsafe_allow_html=True)
            mini_fig = qa_mini_donut(row, row["QA Name"])
            st.plotly_chart(mini_fig, use_container_width=True, config=PLOTLY_CONFIG,
                             key=f"mini_{row['QA Name']}_{period_label}")
            qa_mini_figs.append((row["QA Name"], mini_fig))



st.markdown('<div class="panel-title" style="margin-top:1.2rem;">Daily Log</div>', unsafe_allow_html=True)
st.caption(
    "Every cleaned daily record for the QAs and period selected in the sidebar filters · sortable. "
    "All rows are shown in this one scrollable table (no separate pages) — scroll inside the "
    "table itself, vertically for more rows and horizontally for the Comment column, to see everything."
)

log_df = df_period.copy()

log_df_display = log_df[["Date", "Day", "QA Name", "Billable Hours", "Non-Billable Hours",
                          "Hours Not Worked", "Total Hours", "Comment"]].sort_values("Date", ascending=False)
log_df_display["Date"] = log_df_display["Date"].dt.strftime("%Y-%m-%d")

st.dataframe(
    log_df_display.round(2),
    use_container_width=True,
    hide_index=True,
    height=380,
    column_config={
        "Date": st.column_config.TextColumn("Date", width=95),
        "Day": st.column_config.TextColumn("Day", width=60),
        "QA Name": st.column_config.TextColumn("QA Name", width=150),
        "Billable Hours": st.column_config.NumberColumn("Billable Hours", width=100, format="%.2f"),
        "Non-Billable Hours": st.column_config.NumberColumn("Non-Billable Hours", width=110, format="%.2f"),
        "Hours Not Worked": st.column_config.NumberColumn("Hours Not Worked", width=110, format="%.2f"),
        "Total Hours": st.column_config.NumberColumn("Total Hours", width=100, format="%.2f"),
        # Comment's explicit width, combined with the explicit widths above,
        # sums to well past any realistic rendered table width even with
        # use_container_width=True -- that's what makes the grid's built-in
        # horizontal scroll actually engage (confirmed directly: scrolling
        # the grid horizontally does bring later columns and more Comment
        # text into view). The scrollbar itself has no visible indicator in
        # Streamlit's grid (a real, confirmed platform limitation, not a bug
        # in this app).
        "Comment": st.column_config.TextColumn(
            "Comment", width=500,
            help="Long comments are truncated here — scroll the table sideways, or drag this column's edge to widen it, to read one in full.",
        ),
    },
)
st.caption(f"{len(log_df_display):,} rows total \u00b7 all shown above, scroll to see the rest \u00b7 the table also scrolls sideways for the Comment column, though the scrollbar itself isn't always visible")

with st.expander("\U0001F4CB View Per-QA Summary Table"):
    st.dataframe(
        qa_summary[["QA Name", "Billable Hours", "Non-Billable Hours", "Hours Not Worked",
                     "Total Hours", "Utilization %", "Days Logged"]].round(1),
        use_container_width=True, hide_index=True,
        column_config={
            # "%.1f%%" -> one decimal place followed by a literal "%" sign,
            # e.g. 53.7 -> "53.7%". This only changes how the number is
            # DISPLAYED in this on-screen table; the underlying value stays a
            # plain float, so it still sorts numerically when the user clicks
            # the column header.
            "Utilization %": st.column_config.NumberColumn("Utilization %", format="%.1f%%"),
        },
    )

# ============================================================================
# EXPORT — explicit "Prepare Export" gate.
# Nothing below is computed until the button is pressed, and once the export
# bytes exist they are cached in session_state so simply changing an unrelated
# filter afterwards does NOT silently regenerate/re-run the export.
# ============================================================================
st.markdown('<div class="panel-title" style="margin-top:1.2rem;">Export</div>', unsafe_allow_html=True)
st.caption("Exports always reflect exactly what is on screen right now. Nothing is generated until you click Prepare Export.")

export_summary = qa_summary[["QA Name", "Billable Hours", "Non-Billable Hours", "Hours Not Worked",
                              "Total Hours", "Utilization %", "Days Logged"]].round(2)
export_detail = df_period[["QA Name", "Date", "Day", "Month", "Billable Hours",
                            "Non-Billable Hours", "Hours Not Worked", "Total Hours", "Comment"]].sort_values(["QA Name", "Date"])

# ---- PDF section toggles -------------------------------------------------
# These checkboxes control ONLY the PDF export -- the Excel export always
# contains every section (Dashboard sheet with the Team Overview charts +
# Individual QA Breakdown images, QA Summary sheet, Detail Data sheet)
# regardless of what's checked here, since to_excel_bytes doesn't take a
# sections argument at all. "Team Overview", "Individual QA Breakdown", and
# "Per-QA Summary Table" default to checked (ON); "Daily Log" defaults to
# UNCHECKED so a Prepare Exports click doesn't produce a long log table in
# the PDF unless the person explicitly opts in.
st.markdown('<div class="panel-sub" style="margin-top:8px; margin-bottom:4px;"><b>PDF sections</b> \u2014 choose what to include in the PDF. The Excel export always contains every section regardless of these checkboxes.</div>', unsafe_allow_html=True)

if "pdf_section_team_overview" not in st.session_state:
    st.session_state["pdf_section_team_overview"] = True
if "pdf_section_individual_breakdown" not in st.session_state:
    st.session_state["pdf_section_individual_breakdown"] = True
if "pdf_section_summary_table" not in st.session_state:
    st.session_state["pdf_section_summary_table"] = True
if "pdf_section_daily_log" not in st.session_state:
    st.session_state["pdf_section_daily_log"] = False  # OFF by default, per requirement

cb1, cb2, cb3, cb4 = st.columns(4)
with cb1:
    pdf_include_team_overview = st.checkbox("Team Overview", key="pdf_section_team_overview")
with cb2:
    pdf_include_individual_breakdown = st.checkbox("Individual QA Breakdown", key="pdf_section_individual_breakdown")
with cb3:
    pdf_include_summary_table = st.checkbox("Per-QA Summary Table", key="pdf_section_summary_table")
with cb4:
    pdf_include_daily_log = st.checkbox("Daily Log", key="pdf_section_daily_log")

pdf_sections = {
    "Team Overview": pdf_include_team_overview,
    "Individual QA Breakdown": pdf_include_individual_breakdown,
    "Per-QA Summary Table": pdf_include_summary_table,
    "Daily Log": pdf_include_daily_log,
}

prepare_clicked = st.button("\u2699\ufe0f Prepare Exports", type="primary", use_container_width=False)

if prepare_clicked:
    st.session_state["_run_export_build"] = True

if st.session_state.get("_run_export_build"):
    st.session_state["_run_export_build"] = False

    # Export charts are rendered with matplotlib (see build_export_chart_pngs),
    # which has no Chrome/Chromium/kaleido dependency — it always works, on a
    # laptop, a server, or a frozen .exe. So there is no viability probe and no
    # "install Chrome" step anymore; images_ok is simply True unless an
    # individual figure raises (each is guarded, and any that fails just falls
    # back to that section's data table).
    images_ok = True
    chart_pngs = {}
    mini_pngs = {}
    with st.spinner("Rendering charts for the exports..."):
        try:
            chart_titles_for_render = [title for title, _ in team_chart_figs]
            qa_names_for_render = [qa_name for qa_name, _ in qa_mini_figs]
            chart_pngs, mini_pngs = build_export_chart_pngs(
                export_summary, billable_total, nonbill_total, notworked_total,
                chart_titles_for_render, qa_names_for_render)
        except Exception as render_err:
            images_ok = False
            st.warning(
                f"Chart images could not be rendered ({render_err}). Proceeding "
                "to build the Excel and PDF **without chart images** — all KPI "
                "numbers and every selected section will still be complete."
            )

    try:
        with st.spinner("Building Excel and PDF reports..."):
            chart_titles = [title for title, _ in team_chart_figs]
            qa_names_for_export = [qa_name for qa_name, _ in qa_mini_figs]
            excel_buf, excel_images_failed = to_excel_bytes(
                export_summary, export_detail, kpis, period_label,
                chart_titles, chart_pngs, qa_names_for_export, mini_pngs,
                include_images=images_ok)
            pdf_buf, pdf_images_failed = to_pdf_bytes(
                export_summary, export_detail, kpis, period_label,
                chart_titles, chart_pngs, qa_names_for_export, mini_pngs,
                include_images=images_ok, sections=pdf_sections)
            st.session_state["export_excel_bytes"] = excel_buf
            st.session_state["export_pdf_bytes"] = pdf_buf
            st.session_state["export_period_label"] = period_label
            st.session_state["export_filename_base"] = _build_export_filename_base(period_label, sel_qas, all_qas)
            st.session_state["export_generated_at"] = datetime.now().strftime("%d %b %Y, %H:%M:%S")

        if images_ok and not excel_images_failed and not pdf_images_failed:
            st.success("Exports ready below, with all charts included \u2014 changing filters now will NOT regenerate them until you click Prepare Exports again.")
        else:
            st.info("Exports ready below \u2014 data-complete, but some chart images could not be rendered. Changing filters now will NOT regenerate them until you click Prepare Exports again.")
    except Exception as e:
        st.error(f"Export generation failed: {e}")
        st.session_state.pop("export_excel_bytes", None)
        st.session_state.pop("export_pdf_bytes", None)

if "export_excel_bytes" in st.session_state:
    gen_at = st.session_state.get("export_generated_at", "unknown time")
    st.caption(f"\u2705 These exports were generated at **{gen_at}**. If that's not recent, click Prepare Exports again after fixing any errors above.")
    e1, e2 = st.columns(2)
    fname_base = st.session_state.get(
        "export_filename_base",
        _build_export_filename_base(st.session_state.get("export_period_label", period_label), sel_qas, all_qas),
    )
    with e1:
        st.download_button(
            "\U0001F4E5 Download Excel",
            data=st.session_state["export_excel_bytes"],
            file_name=f"{fname_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with e2:
        st.download_button(
            "\U0001F4C4 Download PDF",
            data=st.session_state["export_pdf_bytes"],
            file_name=f"{fname_base}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
